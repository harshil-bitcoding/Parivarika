import os
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from parivar.models import Person, ParentChildRelation, TranslatePerson, Surname

class Command(BaseCommand):
    help = 'Export all main Person data to a single Excel file with multiple sheets by surname'

    def handle(self, *args, **options):
        # Create workbook
        wb = Workbook()
        
        # 1. Create Dashbord sheet (default sheet renamed)
        dashbord = wb.active
        dashbord.title = "Dashbord"
        
        # 2. Create Dummy sheet
        wb.create_sheet("Dummy")
        
        # Get all unique surname names that have active persons
        unique_surname_names = Person.objects.filter(
            is_deleted=False, 
            surname__isnull=False
        ).values_list('surname__name', flat=True).distinct().order_by('surname__name')
        
        self.stdout.write(f"Found {len(unique_surname_names)} unique surname names. Starting export...")

        # Common styles
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for surname_name in unique_surname_names:
            if not surname_name:
                continue
                
            # Create a sheet for the surname
            # openpyxl limits sheet titles to 31 characters
            ws = wb.create_sheet(title=surname_name[:31])
            
            # Row 1: Categories (with merging)
            ws.merge_cells('A1:B1')
            ws['A1'] = "Firstname"
            
            ws.merge_cells('C1:D1')
            ws['C1'] = "Father name"
            
            ws['E1'] = "Column7"
            ws['F1'] = "Surname"
            ws['G1'] = "Birt Date"
            
            ws.merge_cells('H1:I1')
            ws['H1'] = "Mobile Number"
            
            ws.merge_cells('J1:K1')
            ws['J1'] = "Out of India ?"
            
            ws.merge_cells('L1:M1')
            ws['L1'] = "Link"

            ws.merge_cells('N1:O1')
            ws['N1'] = "Image"

            # Apply styling to Row 1
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # Row 2: Sub-headers
            headers_v2 = [
                "In English", "In Gujarati", "In English", "In Gujarati", 
                "Full name", "Surname", "In DD-MM-YY", "Main", "Optional", 
                "Country Name", "Mobile Number", "Name of Father", "Name of Son",
                "Profile", "Thumb profile"
            ]
            for col_num, header in enumerate(headers_v2, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # Fetch persons for this surname name
            persons = Person.objects.filter(surname__name=surname_name, is_deleted=False)
            
            row_idx = 3
            base_url = getattr(settings, 'BASE_URL', '').rstrip('/')
            for person in persons:
                # Look up relations
                father_relation = ParentChildRelation.objects.filter(child=person, is_deleted=False).first()
                son_relation = ParentChildRelation.objects.filter(parent=person, is_deleted=False).first()
                
                # Look up Gujarati translations
                translation = TranslatePerson.objects.filter(person_id=person, language='guj', is_deleted=False).first()
                
                # Professional Date Formatting
                dob = person.date_of_birth or ''
                if dob:
                    clean_dob = dob.split(' ')[0]
                    if '-' in clean_dob:
                        parts = clean_dob.split('-')
                        if len(parts) == 3 and len(parts[0]) == 4:
                            dob = f"{parts[2]}/{parts[1]}/{parts[0]}"
                        else:
                            dob = clean_dob
                    else:
                        dob = clean_dob
                
                if '00:00:00' in dob or '0000-00-00' in dob:
                    dob = ''

                full_name = f"{person.first_name or ''} {person.middle_name or ''}".strip()

                row_data = [
                    person.first_name or '',
                    translation.first_name if translation else '',
                    person.middle_name or '',
                    translation.middle_name if translation else '',
                    full_name,
                    surname_name,
                    dob,
                    person.mobile_number1 or '',
                    person.mobile_number2 or '',
                    (person.out_of_country.name if person.out_of_country and person.out_of_country.name.lower() != 'india' else ''),
                    person.out_of_mobile or '',
                    (f"{father_relation.parent.first_name} {father_relation.parent.middle_name or ''}".strip() if father_relation and father_relation.parent else ''),
                    (f"{son_relation.child.first_name} {son_relation.child.middle_name or ''}".strip() if son_relation and son_relation.child else ''),
                    person.profile.url if person.profile else '',
                    person.thumb_profile.url if person.thumb_profile else ''
                    # (base_url + person.profile.url) if person.profile else '',
                    # (base_url + person.thumb_profile.url) if person.thumb_profile else ''
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.value = value
                    cell.border = thin_border
                
                row_idx += 1

        # Save the workbook
        output_filename = "Parivarbook_Bila.xlsx"
        wb.save(output_filename)
        self.stdout.write(self.style.SUCCESS(f"Successfully exported all data to {output_filename}"))

