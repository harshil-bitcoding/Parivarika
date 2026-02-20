
import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter

def clean_text(text):
    """Normalize text: strip whitespace and convert to lowercase."""
    if not text:
        return ""
    return str(text).strip().lower()

def load_reference_data(xlsx_path):
    """
    Loads profile data from the exported Excel file.
    Returns a dictionary: 
    { 
      sheet_name: { 
          (clean_first, clean_middle): (profile_url, thumb_url) 
      } 
    }
    """
    print(f"Loading reference data from {xlsx_path}...")
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    except FileNotFoundError:
        print(f"Error: file {xlsx_path} not found.")
        sys.exit(1)

    lookup = {}
    total_records = 0
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ['Dashbord', 'Dummy']:
            continue
            
        ws = wb[sheet_name]
        sheet_lookup = {}
        
        # Iterate rows, skipping header (rows 1 and 2)
        # Using iter_rows with values_only=True
        # Row 3 is start of data
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row: 
                continue
                
            # Column indices (0-based tuple from values_only):
            # A=0 (Firstname In English)
            # C=2 (Father name In English)
            # N=13 (Profile)
            # O=14 (Thumb Profile)
            
            try:
                # Ensure row has enough columns
                if len(row) < 15:
                    continue

                first = row[0]
                middle = row[2]
                profile = row[13] # Col N
                thumb = row[14]   # Col O
                
                if profile or thumb:
                    key = (clean_text(first), clean_text(middle))
                    sheet_lookup[key] = (profile, thumb)
                    total_records += 1
            except IndexError:
                continue
        
        if sheet_lookup:
            lookup[sheet_name] = sheet_lookup
                
    print(f"Loaded {total_records} records across {len(lookup)} sheets.")
    return lookup

def process_target_excel(target_xlsx_path, lookup):
    """
    Reads target Excel, matches records BY SHEET NAME, and saves a new Excel file.
    """
    output_xlsx_path = "export_profile.xlsx" # As requested by user
    print(f"Processing Excel {target_xlsx_path} -> {output_xlsx_path}...")
    
    try:
        wb = openpyxl.load_workbook(target_xlsx_path)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    matched_count = 0
    
    for sheet_name in wb.sheetnames:
        if sheet_name not in lookup:
            # print(f"Skipping sheet '{sheet_name}' (no matching source data).")
            continue
            
        print(f"Processing sheet '{sheet_name}'...")
        ws = wb[sheet_name]
        sheet_data = lookup[sheet_name]
        
        # Target Structure Analysis based on screenshots:
        # Row 1: Merged Headers
        # Row 2: Sub Headers
        # Data starts Row 3
        # Columns:
        # A (1): Firstname (In English)
        # C (3): Father name (In English)
        # N (14): Profile
        # O (15): Thumb profile
        
        # We will write to N and O.
        # Check heuristics? User said structure is key. 
        # Let's verify headers briefly.
        
        # Check Row 2, Col 14/15
        val_n2 = ws.cell(row=2, column=14).value
        val_o2 = ws.cell(row=2, column=15).value
        
        # If headers are missing/different, warn but proceed?
        # User said "upload profile and thumb profile" implying they might be empty or missing.
        # Let's ensure headers exist.
        if sheet_name == "Thummar": # Just checking one known sheet for debug print
             pass
             
        ws.cell(row=2, column=14).value = "Profile" 
        ws.cell(row=2, column=15).value = "Thumb profile"
        
        # Merge N1:O1 if not merged? User's screenshot shows "Image" there.
        # We can set N1 to "Image"
        ws.cell(row=1, column=14).value = "Image"
        # Merging cells logic omitted to avoid breaking existing merges unless necessary
        
        for row_idx in range(3, ws.max_row + 1):
            # Read names
            first_val = ws.cell(row=row_idx, column=1).value
            middle_val = ws.cell(row=row_idx, column=3).value
            
            key = (clean_text(first_val), clean_text(middle_val))
            
            if key in sheet_data:
                profile_val, thumb_val = sheet_data[key]
                if profile_val:
                    ws.cell(row=row_idx, column=14).value = profile_val
                if thumb_val:
                    ws.cell(row=row_idx, column=15).value = thumb_val
                matched_count += 1
                
    wb.save(output_xlsx_path)
    print(f"Done. Matched {matched_count} records.")
    print(f"Output saved to: {output_xlsx_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python merge_profiles.py <target_xlsx_path> [reference_xlsx_path]")
        sys.exit(1)
        
    target_file = sys.argv[1]
    
    # Default reference path
    ref_xlsx = "Parivarbook_Bila.xlsx"
    if len(sys.argv) > 2:
        ref_xlsx = sys.argv[2]
        
    if not os.path.exists(target_file):
        print(f"Error: Target file {target_file} not found.")
        sys.exit(1)
        
    lookup_data = load_reference_data(ref_xlsx)
    process_target_excel(target_file, lookup_data)
